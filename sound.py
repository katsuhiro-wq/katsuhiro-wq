# 周波数解析に使用したライブラリ
import numpy as np
import matplotlib.pyplot as plt
import soundfile as sf
import openpyxl
from scipy import integrate
from matplotlib.font_manager import FontProperties
fp = FontProperties(fname=r'/System/Library/Fonts/ヒラギノ角ゴシック W6.ttc', size=16)

#メイン関数
def FFT_main(t, x, dt):
    sp_data = data_sp(t, x)
    FFT_list = []
    for data_count in sp_data:
        FFT_count = FFT(data_count, dt)
        FFT_list.append(FFT_count)
    fq_ave = FFT_list[0][0]
    PS_ave = np.zeros(len(fq_ave))
    for i in range(len(FFT_list)):
        PS_ave = PS_ave + FFT_list[i][1]
    PS_ave = PS_ave/(i+1)
    plot_FFT(t, x, fq_ave, PS_ave)
    return fq_ave, PS_ave

#シャント音データのオーバーラップ処理を行う関数
def data_sp(t ,x):
    sp_data = []
    frame = int(len(t) * 0.1)
    overlap = int(frame * 0.5)
    start = 0
    end = start + frame
    while True:
        t_cont = t[start:end]
        x_cont = x[start:end]
        sp_data.append([t_cont, x_cont])
        start = start + (frame - overlap)
        end = start + frame
        if end > len(t):
            break
    return np.array(sp_data)

#窓関数をかけた後にフーリエ変換を行う関数
def FFT(data_input, dt):
    N = len(data_input[0])
    window = data_input[1] * np.hanning(N)
    F = np.fft.fft(window)
    F_abs = np.abs(F)
    AS = F_abs / N * 2
    PS = AS ** 2
    fq = np.linspace(0, 1.0 / dt, N)
    PS = 1/(sum(np.hanning(N)) / N) * PS
    fq = fq[:int(N/2)+1]
    PS = PS[:int(N/2)+1]
    return [fq, PS]

#フーリエ変換後のPSをグラフに抽出する関数
def plot_FFT(t, x, fq, PS):
    fig = plt.figure(figsize=(16, 6))
    ax2 = fig.add_subplot(121)
    plt.plot(t, x)
    plt.xlabel("時間[s]",fontproperties=fp)
    plt.ylabel("振幅",fontproperties=fp)
    plt.title("時間軸",fontproperties=fp)
    ax2 = fig.add_subplot(122)
    plt.plot(fq, PS)
    plt.xlabel('周波数[Hz]',fontproperties=fp)
    plt.ylabel('パワースペクトル',fontproperties=fp)
    plt.title('周波数軸',fontproperties=fp)
    plt.savefig('va_fft.png', dpi=300)
    plt.show()
    PS_simp = PS[:-1]
    fq_simp = fq[:-1]

#フーリエ変換で得られた結果をExcelファイルに保存するプログラム
    start, end = 0, 100
    wb = openpyxl.Workbook()
    ws = wb.active
    for i in range(0, 30):
        y_simp = PS_simp[start:end]
        x_simp = fq_simp[start:end]
        simp = integrate.simps(y_simp,x_simp)
        ws.cell(row=i+1,column=1,value = str(int(start/0.5))+"~"+str(int(end/0.5))+"Hz")
        ws.cell(row=i+1,column=2,value = simp)
        start += 100
        end += 100
    wb.save("vasound.xlsx")
    return 0

#Sound Fileを用いたWAVファイルの読み込みとフーリエ変換のメイン関数を呼び出す関数
if __name__ == "__main__":
    path = "竹鼻立原町 17.wav"
    x, sr = sf.read(path)
    n = len(x)
    t = np.linspace(0, n / sr ,n)
    dt = 1 / sr
    FFT_main(t, x, dt)
  